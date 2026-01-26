"""
exporter.py - 多维数据提取并导出为带超链接的 Excel 报告

核心职责:
1. 从 SQLite 提取数据
2. 使用 pandas + openpyxl 导出 Excel
3. URL 字段自动转换为可点击超链接
"""

import os
import sqlite3
from datetime import datetime
from typing import List, Optional, Union
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .storage_manager import StorageManager


class Exporter:
    """数据导出器：将数据库内容导出为投研标准 Excel"""
    
    def __init__(self, storage_manager: StorageManager = None, output_dir: str = None):
        self.sm = storage_manager or StorageManager()
        
        if output_dir is None:
            output_dir = Path(__file__).parent.parent / "exports"
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_excel(
        self,
        author: Union[str, List[str], None] = None,
        start_date: str = None,
        end_date: str = None,
        keyword: str = None,
        filename: str = None,
        external_data: List[dict] = None
    ) -> str:
        """
        导出数据到 Excel 文件
        
        Args:
            author: 按作者筛选
            start_date: 起始日期 (YYYY-MM-DD)
            end_date: 结束日期 (YYYY-MM-DD)
            keyword: 全文搜索关键词
            filename: 自定义文件名，默认自动生成
            external_data: 可选，直接传入要导出的数据列表（如果提供则跳过数据库查询）
            
        Returns:
            生成的 Excel 文件路径
        """
        # 1. 获取数据
        if external_data is not None:
            tweets = external_data
        else:
            tweets = self.sm.get_tweets(
                author=author,
                start_date=start_date,
                end_date=end_date,
                keyword=keyword
            )
        
        if not tweets:
            print("⚠️ 没有找到符合条件的数据")
            return None
        
        # 2. 转换为 DataFrame
        # 直接使用数据库返回的列
        df = pd.DataFrame(tweets)
        
        # 选择并重命名需要的列
        columns_mapping = {
            'author': '作者',
            # 'author_name': '作者昵称', # 数据库可能还没存 nick name，或者在 content 表里没有
            'text': '内容',
            'publish_time': '发布时间',
            'url': '原文链接',
            'like_count': '点赞数',
            'retweet_count': '转发数',
            'reply_count': '评论数',
            'quote_count': '引用数',
            'view_count': '阅读量',
            'lang': '语言',
            'author_followers': '作者粉丝数',
            'platform': '平台',
            'is_retweet': '是否转发'
        }
        
        # 确保 DataFrame 包含所有必要的列 (防止全空时报错)
        for col in columns_mapping.keys():
            if col not in df.columns:
                df[col] = None # 或默认值
        
        # 只保留存在的列，并按顺序排列（如果存在）
        preferred_order = [
            '作者', '内容', '发布时间', 
            '点赞数', '评论数', '转发数', '引用数', '阅读量', '作者粉丝数',
            '原文链接', '语言', '是否转发', '平台'
        ]
        
        # 重命名
        df.rename(columns=columns_mapping, inplace=True)
        
        # 筛选存在的列
        final_cols = [col for col in preferred_order if col in df.columns]
        
        # 添加未在 preferred_order 中但存在于 mapping 后的列
        mapped_cols = list(columns_mapping.values())
        remaining_mapped = [col for col in df.columns if col in mapped_cols and col not in final_cols]
        final_cols.extend(remaining_mapped)
        
        # ✅ 关键修复：添加所有未映射的列（即动态标注字段）
        extra_cols = [col for col in df.columns if col not in final_cols]
        final_cols.extend(extra_cols)
        
        df = df[final_cols].copy()
        
        # 处理是否转发列
        if '是否转发' in df.columns:
            df['是否转发'] = df['是否转发'].apply(lambda x: '是' if x else '否')
        
        # 3. 生成文件名：YYYYMMDD_HHMMSS_作者_数据导出.xlsx
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if isinstance(author, list):
                if len(author) > 3:
                    author_part = "多博主_"
                else:
                    author_part = f"{'_'.join(author)}_"
            else:
                author_part = f"{author}_" if author else "全部_"
            filename = f"{timestamp}_{author_part}数据导出.xlsx"
        
        filepath = self.output_dir / filename
        
        # 4. 导出到 Excel（临时，不带超链接）
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        # 5. 使用 openpyxl 添加超链接
        self._add_hyperlinks(filepath, df)
        
        print(f"✅ 数据已导出: {filepath}")
        print(f"   共 {len(df)} 条记录")
        
        return str(filepath)
    
    def _add_hyperlinks(self, filepath: str, df: pd.DataFrame):
        """
        为 Excel 中的 URL 列添加可点击超链接
        
        Args:
            filepath: Excel 文件路径
            df: 原始 DataFrame，用于定位 URL 列
        """
        wb = load_workbook(filepath)
        ws = wb.active
        
        # 找到"原文链接"列的位置
        url_col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == '原文链接':
                url_col_idx = idx
                break
        
        if url_col_idx is None:
            wb.save(filepath)
            return
        
        col_letter = get_column_letter(url_col_idx)
        
        # 从第 2 行开始（跳过表头）添加超链接
        for row_idx in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row_idx}"]
            url = cell.value
            
            if url and isinstance(url, str) and url.startswith('http'):
                cell.hyperlink = url
                cell.style = 'Hyperlink'
                # 显示为更友好的文本
                cell.value = "🔗 查看原文"
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # 限制最大宽度
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
    
    def export_summary(
        self,
        author: str = None,
        start_date: str = None,
        end_date: str = None
    ) -> dict:
        """
        生成数据摘要统计
        
        Returns:
            包含统计信息的字典
        """
        tweets = self.sm.get_tweets(
            author=author,
            start_date=start_date,
            end_date=end_date
        )
        
        if not tweets:
            return {"total": 0}
        
        df = pd.DataFrame(tweets)
        
        summary = {
            "total": len(df),
            "authors": df['author'].nunique() if 'author' in df.columns else 0,
            "date_range": {
                "start": df['publish_time'].min() if 'publish_time' in df.columns else None,
                "end": df['publish_time'].max() if 'publish_time' in df.columns else None
            },
            "retweets": int(df['is_retweet'].sum()) if 'is_retweet' in df.columns else 0,
            "original": len(df) - int(df['is_retweet'].sum()) if 'is_retweet' in df.columns else len(df)
        }
        
        # 按作者统计
        if 'author' in df.columns:
            author_counts = df['author'].value_counts().to_dict()
            summary["by_author"] = dict(list(author_counts.items())[:10])  # 只取前10
        
        return summary
    
    def export_annotated_tweets(
        self,
        author: str = None,
        filename: str = None
    ) -> str:
        """
        导出已标注的推文数据
        
        Args:
            author: 可选，只导出特定作者
            filename: 自定义文件名
            
        Returns:
            导出文件路径
        """
        # 获取已标注数据
        conn = sqlite3.connect(self.sm.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        query = "SELECT * FROM content WHERE annotated_at IS NOT NULL"
        params = []
        
        if author:
            if isinstance(author, list):
                if author:
                    placeholders = ', '.join(['?'] * len(author))
                    query += f" AND author IN ({placeholders})"
                    params.extend(author)
            else:
                query += " AND author = ?"
                params.append(author)
        
        query += " ORDER BY annotated_at DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        
        if not rows:
            print("⚠️ 没有已标注的数据")
            return None
        
        tweets = [dict(row) for row in rows]
        
        # 转换为 DataFrame
        df = pd.DataFrame(tweets)
        
        # 列映射（包含标注字段）
        columns_mapping = {
            'author': '作者',
            'text': '内容',
            'publish_time': '发布时间',
            
            # 标注字段
            'sentiment': '情感倾向',
            'topic_category': '主题分类',
            'importance_score': '重要性',
            'keywords': '关键词',
            
            # 互动数据
            'like_count': '点赞数',
            'retweet_count': '转发数',
            'reply_count': '评论数',
            'view_count': '阅读量',
            
            'url': '原文链接',
            'annotated_at': '标注时间',
            'lang': '语言'
        }
        
        # 确保列存在
        for col in columns_mapping.keys():
            if col not in df.columns:
                df[col] = None
        
        # 重命名
        df.rename(columns=columns_mapping, inplace=True)
        
        # 排序
        preferred_order = [
            '作者', '内容', '发布时间',
            '情感倾向', '主题分类', '重要性', '关键词',
            '点赞数', '转发数', '评论数', '阅读量',
            '原文链接', '标注时间', '语言'
        ]
        
        final_cols = [col for col in preferred_order if col in df.columns]
        df = df[final_cols].copy()
        
        # 生成文件名：日期_作者_已标注数据
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if isinstance(author, list):
                if len(author) > 3:
                    author_part = "多博主_"
                else:
                    author_part = f"{'_'.join(author)}_"
            else:
                author_part = f"{author}_" if author else "全部_"
            filename = f"{timestamp}_{author_part}已标注数据.xlsx"
        
        filepath = self.output_dir / filename
        
        # 导出
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        # 添加超链接
        self._add_hyperlinks(filepath, df)
        
        print(f"✅ 数据已导出: {filepath}")
        print(f"   共 {len(df)} 条已标注记录")
        
        return str(filepath)


# ==================== 测试代码 ====================
if __name__ == "__main__":
    from storage_manager import StorageManager
    
    sm = StorageManager()
    
    # 添加测试数据
    test_tweets = [
        {
            "tweet_id": "test_001",
            "author": "test_user",
            "text": "这是一条测试推文",
            "created_at": "2024-01-15T10:30:00",
            "url": "https://x.com/test_user/status/test_001",
            "is_retweet": False
        },
        {
            "tweet_id": "test_002",
            "author": "test_user",
            "text": "这是第二条测试推文",
            "created_at": "2024-01-16T14:20:00",
            "url": "https://x.com/test_user/status/test_002",
            "is_retweet": False
        }
    ]
    
    sm.save_tweets(test_tweets)
    
    # 导出测试
    exporter = Exporter(sm)
    filepath = exporter.export_to_excel(author="test_user")
    
    # 摘要测试
    summary = exporter.export_summary()
    print(f"数据摘要: {summary}")
